import openpyxl
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
import os

# Function to search for images using Bing Image Search API
def search_image(query):
    subscription_key = "YOUR_BING_API_KEY"
    search_url = "https://api.bing.microsoft.com/v7.0/images/search"
    headers = {"Ocp-Apim-Subscription-Key": subscription_key}
    params = {"q": query, "license": "public", "imageType": "photo"}
    response = requests.get(search_url, headers=headers, params=params)
    response.raise_for_status()
    search_results = response.json()
    return search_results['value'][0]['contentUrl']

# Load an existing Excel file
wb = openpyxl.load_workbook('your_excel_file.xlsx')
sheet = wb.active

# Assuming the codes are in the first column, starting from row 1
for row in range(1, sheet.max_row + 1):
    code = sheet.cell(row=row, column=1).value
    if code:
        try:
            image_url = search_image(code)
            image_response = requests.get(image_url)
            image_response.raise_for_status()
            image_bytes = BytesIO(image_response.content)
            img = Image(image_bytes)
            # Assuming you want to add the image in the second column
            img_cell = 'B' + str(row)
            sheet.add_image(img, img_cell)
        except Exception as e:
            print(f"Failed to retrieve or add image for {code}: {str(e)}")

# Save the updated workbook
wb.save('updated_excel_file.xlsx')
