#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%      first ex        %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

# from openpyxl import workbook, load_workbook

# # first you nedd close excel if you don't you face on error
# # also if you have excel file use load_workbook if you don't use workbook()
# wb = load_workbook("/Users/adam/Desktop/scrach/python/job/one_for_all/learn/test.xlsx") # path of excel file
# ws = wb.active # open active sheet  ws = work sheet
# print(ws)   # name of active sheet
# ws = wb['shex'] # open spicfic sheet 
# print(ws)   # name of that sheet you mention for now the sheet name is shex
# ws = wb.create_sheet('test') # create new sheet and assign to ws wich mean test sheet is active sheet
# print(wb.sheetnames) # print all sheet that i have in the file
# print(ws['A1']) # name of cell <Cell 'adam'.A1>
# print(ws['A1'].value) # value of the cell
# ws['A2'] = "adam" # change value of the cell
# ws['A3'].value = "san" # also you can do this no differtent
# # if you don't write wb.save(path of file) you can't see changes
# wb.save("/Users/adam/Desktop/scrach/python/job/one_for_all/learn/test.xlsx") 




#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%      second ex        %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

# # warring W in workbook import is capital
# from openpyxl import  Workbook,load_workbook

# # create new work book and new sheet also name it ans some things
# wb = Workbook() # create new workbook
# ws = wb.active # open active book
# ws.title = 'firstSheet' # change name of worksheet to firstsheet

# ws.append(['name','age','mony','salary']) # adding this itme to first row
# ws.append(['name','age','mony','salary']) # adding this itme to second row
# wb.save("temp.xlsx")


#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%      third ex        %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

from openpyxl import Workbook,load_workbook


